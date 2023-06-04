from copy import copy

import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill


class ExcelManager:
    def __init__(self):
        super()

    @staticmethod
    def new_wb():
        return openpyxl.Workbook()

    @staticmethod
    def update_spreadsheet(wb: openpyxl.Workbook, df, start_col: int, start_row: int, sheet_name: str, progress):
        redFill = PatternFill(start_color='FFFF0000',
                              end_color='FFFF0000',
                              fill_type='solid')
        if isinstance(df, pd.DataFrame):
            for ir in range(0, len(df)):
                for ic in range(0, len(df.iloc[ir])):
                    wb[sheet_name].cell(start_row + ir, start_col + ic).value = df.iloc[ir][ic]
        if isinstance(df, list):
            shift = 100 / len(df)
            for i in range(3, len(df)):

                progress.setValue(i * shift)
                progress.setFormat('Запись новых значений {}%'.format(i * shift))

                tmp = str(df[i][0])
                flag = str(df[i][1])
                if tmp == 'nan':
                    tmp = ''
                wb[sheet_name].cell(i + 1, start_col).value = tmp
                if flag == '1':
                    wb[sheet_name].cell(i + 1, start_col).fill = redFill
        return wb

    @staticmethod
    def load_excel(path: str):
        return openpyxl.load_workbook(path, keep_links=False)

    @staticmethod
    def save_excel(wb, path):
        return wb.save(path)

    @staticmethod
    def process_unmerge_cells(ws, merged=None, coords=None, col=None, progress=None):
        if merged:
            shift = 100 / len(merged)
            for index, group in enumerate(list(merged)):
                progress.setValue(index * shift)
                progress.setFormat('Подготовка excel к записи {}%'.format(index * shift))
                min_col, min_row, max_col, max_row = group.bounds
                if min_col != col and min_row != max_row:
                    continue
                cell_start = ws.cell(row=min_row, column=min_col)
                top_left_cell_value = cell_start.value
                ws.unmerge_cells(str(group))
                for i_row in range(min_row, max_row + 1):
                    for j_col in range(min_col, max_col + 1):
                        ws.cell(row=i_row, column=j_col, value=top_left_cell_value)
                        ws.cell(row=i_row, column=j_col).alignment = copy(cell_start.alignment)
                        ws.cell(row=i_row, column=j_col).border = copy(cell_start.border)
                        ws.cell(row=i_row, column=j_col).font = copy(cell_start.font)
                        ws.cell(row=i_row, column=j_col).fill = copy(cell_start.fill)
            return ws
        if coords:
            start_row, start_column, end_row, end_column = coords
            cell_start = ws.cell(row=start_row, column=start_column)
            top_left_cell_value = cell_start.value
            ws.unmerge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column)
            for i_row in range(start_row, end_row + 1):
                for j_col in range(start_column, end_column + 1):
                    ws.cell(row=i_row, column=j_col, value=top_left_cell_value)
                    ws.cell(row=i_row, column=j_col).alignment = copy(cell_start.alignment)
                    ws.cell(row=i_row, column=j_col).border = copy(cell_start.border)
                    ws.cell(row=i_row, column=j_col).font = copy(cell_start.font)
                    ws.cell(row=i_row, column=j_col).fill = copy(cell_start.fill)

        return ws

    @staticmethod
    def process_merge_cells(ws, merged=None, coords=None, col=None, progress=None):
        if merged:
            shift = 100 / len(merged)
            for index, group in enumerate(list(merged)):
                progress.setValue(index * shift)
                progress.setFormat('Форматирование excel {}%'.format(index * shift))

                min_col, min_row, max_col, max_row = group.bounds
                if min_col != col and min_row != max_row:
                    continue
                ws.merge_cells(str(group))
            return ws
        if coords:
            start_row, start_column, end_row, end_column = coords
            ws.merge_cells(start_row=start_row, start_column=start_column,
                           end_row=end_row, end_column=end_column)

        return ws
