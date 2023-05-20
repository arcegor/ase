import copy
import sys
from pathlib import Path

import numpy as np
from PyQt5 import QtWidgets
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import QFileDialog, QMessageBox
from openpyxl import load_workbook
from openpyxl import Workbook
from ui import Ui_Upload
from FileManager import FileManager
import pandas as pd


class MainWindow(QtWidgets.QMainWindow):

    def __init__(self):
        super().__init__()
        self.ui = Ui_Upload()
        self.ui.setupUi(self)

        # Здесь можно настроить html разметку для приветствия
        self.html = '<h3>Добро пожаловать!</h3>'

        # Кнопки
        self.ui.buttonChooseSource.clicked.connect(self.buttonChooseSource_clicked)
        self.ui.buttonFindCollisions.clicked.connect(self.buttonFindCollisions_clicked)
        self.ui.buttonDownloadFixed.clicked.connect(self.buttonDownloadFixed_clicked)
        self.ui.buttonDownloadReport.clicked.connect(self.buttonDownloadReport_clicked)
        self.ui.buttonChooseTarget.clicked.connect(self.buttonChooseTarget_clicked)
        self.ui.buttonReset.clicked.connect(self.buttonReset_clicked)
        self.ui.buttonFindCollisions.setDisabled(True)
        self.ui.buttonDownloadFixed.setDisabled(True)
        self.ui.buttonDownloadReport.setDisabled(True)
        self.ui.buttonChooseTarget.setDisabled(True)
        self.fm = FileManager()

        self.ui.SomeInfo.setHtml(self.html)

    def buttonChooseSource_clicked(self):
        filename, _ = QFileDialog.getOpenFileName(self, 'Open file', './~', '(*.xls *.xlsx *.xlsm)')
        if not filename:
            return
        wb = self.fm.em.load_excel(filename)
        df = pd.DataFrame(wb.active.values)
        ws = wb.active
        merged = copy.deepcopy(ws.merged_cells.ranges)
        sn = ws.title
        ws_new = self.fm.em.process_unmerge_cells(ws, merged)
        ws_new = self.fm.em.process_merge_cells(ws_new, merged)
        wb.active = ws_new

        #wb_new = self.fm.em.update_spreadsheet(wb, df, 1, 1, sheet_name=sn)
        flag = self.fm.em.save_excel(wb, '12345.xlsx')
        if self.is_valid():
            self.ui.buttonChooseTarget.setEnabled(True)
            self.ui.buttonChooseSource.setDisabled(True)
            self.ui.SomeInfo.append(f'Эталонный файл {filename} успешно загружен!\nВыберите проверяемый файл\n')

    def buttonChooseTarget_clicked(self):
        filename, _ = QFileDialog.getOpenFileName(self, 'Open file', './~', '(*.xls *.xlsx *.xlsm)')
        if not filename:
            return
        self.fm.data['target'] = pd.read_excel(filename)
        if self.is_valid():
            self.ui.buttonChooseTarget.setDisabled(True)
            self.ui.SomeInfo.append(f'Проверяемый файл {filename} успешно загружен!\nТеперь можно провести проверку!\n')
            self.ui.buttonFindCollisions.setEnabled(True)

    def buttonFindCollisions_clicked(self):
        if not self.fm.process():
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText("Ошибка обработки!")
            msg.setWindowTitle("Ошибка!")
            msg.exec_()
            return
        self.ui.buttonDownloadReport.setEnabled(True)
        self.ui.buttonDownloadFixed.setEnabled(True)
        self.ui.buttonFindCollisions.setDisabled(True)

    def buttonDownloadFixed_clicked(self):
        if self.fm.get_result('target'):
            self.ui.SomeInfo.setText('\n'.join(str(x) for x in self.fm.result.keys()))
            self.ui.buttonDownloadFixed.setDisabled(True)

    def buttonDownloadReport_clicked(self):
        if self.fm.get_result('source'):
            self.ui.SomeInfo.setText('\n'.join(str(x) for x in self.fm.result.keys()))
            self.ui.buttonDownloadFixed.setDisabled(True)

    def is_valid(self):
        if not self.fm.validate():
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText("Документы не соответствуют шаблону!")
            msg.setWindowTitle("Ошибка валидации!")
            msg.exec_()
            return False
        return True

    def buttonReset_clicked(self):
        self.fm.clear()
        self.ui.buttonDownloadFixed.setDisabled(True)
        self.ui.buttonDownloadFixed.setDisabled(True)
        self.ui.buttonFindCollisions.setDisabled(True)
        self.ui.buttonChooseSource.setEnabled(True)
        self.ui.buttonChooseTarget.setDisabled(True)
        self.ui.SomeInfo.setHtml(self.html)


if __name__ == '__main__':
    app = QtWidgets.QApplication([])
    app.setWindowIcon(QIcon('atom.ico'))
    app.setStyle('Fusion')
    application = MainWindow()
    application.show()
    sys.exit(app.exec())
